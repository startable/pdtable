"""Machinery to read/write Tables in an Excel workbook using openpyxl as engine."""
from itertools import chain
from os import PathLike
from typing import Union, Iterable, Sequence, Any, Dict, List, Tuple, Optional

import openpyxl

try:
    from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet
except ImportError:
    # openpyxl < 2.6
    from openpyxl.worksheet import Worksheet as OpenpyxlWorksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill, Color, Alignment
from openpyxl.utils import get_column_letter

from pdtable import Table
from pdtable.io._represent import _represent_row_elements, _represent_col_elements

DEFAULT_SHEET_NAME = "Sheet1"
DEFAULT_STYLE_SPEC = {
    "table_name": {
        "font": {"color": "1F4E78", "bold": True,},  # hex color code
        "fill": {"color": "D9D9D9",},  # RGB color code
    },
    "destinations": {"font": {"color": "808080", "bold": True,}, "fill": {"color": "D9D9D9",},},
    "column_names": {"fill": {"color": "F2F2F2",}, "font": {"bold": True,},},
    "units": {"fill": {"color": "F2F2F2",},},
    "values": {},
}


def read_cell_rows_openpyxl(path: Union[str, PathLike]) -> Iterable[Sequence[Any]]:
    """Reads from an Excel workbook, yielding one row of cells at a time."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    for ws in wb.worksheets:
        yield from ws.iter_rows(values_only=True)


def write_excel_openpyxl(tables, path, na_rep, styles, sep_lines):
    """Write tables to an Excel workbook at the specified path."""

    if not isinstance(tables, Dict):
        # For convenience, pack it in a dict
        tables = {DEFAULT_SHEET_NAME: tables}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the one sheet that openpyxl creates by default

    for sheet_name in tables:

        tabs = tables[sheet_name]
        if isinstance(tabs, Table):
            # For convenience, pack single table in an iterable
            tabs = [tabs]

        table_dimensions = []
        ws = wb.create_sheet(title=sheet_name)
        for t in tabs:
            # Keep track of table dimensions for formatting as tuples (num rows, num cols, transposed)
            table_dimensions.append((len(t.df), len(t.df.columns), t.metadata.transposed))
            _append_table_to_openpyxl_worksheet(t, ws, sep_lines, na_rep)

        if styles:
            styles = DEFAULT_STYLE_SPEC if styles is True else styles
            _style_tables_in_worksheet(ws, table_dimensions, styles, sep_lines)

    wb.save(path)


def _append_table_to_openpyxl_worksheet(
    table: Table, ws: OpenpyxlWorksheet, sep_lines: int, na_rep: str = "-"
) -> None:
    """Write table at end of sheet, leaving sep_lines blank lines before."""
    units = table.units
    if table.metadata.transposed:
        ws.append([f"**{table.name}*"])
        ws.append([" ".join(str(x) for x in table.metadata.destinations)])
        for col in table:
            ws.append(
                [str(col.name), str(col.unit)]
                + list(_represent_col_elements(col.values, col.unit, na_rep)),
            )
    else:
        ws.append([f"**{table.name}"])
        ws.append([" ".join(str(x) for x in table.metadata.destinations)])
        ws.append(table.column_names)
        ws.append(units)
        for row in table.df.itertuples(index=False, name=None):
            # TODO: apply format string specified in ColumnMetadata
            ws.append(_represent_row_elements(row, units, na_rep))

    for _ in range(sep_lines):
        ws.append([])  # blank row marking table end


def deep_get(dictionary, keys, default=None):
    """Get value from nested dictionaries.

    Modified from: https://stackoverflow.com/a/50173148/119775

    Example:
        d = {'meta': {'status': 'OK', 'status_code': 200}}
        deep_get(d, ['meta', 'status_code'])          # => 200
        deep_get(d, ['garbage', 'status_code'])       # => None
        deep_get(d, ['meta', 'garbage'], default='-') # => '-'
    """
    if dictionary is None:
        return default
    if not keys:
        return dictionary
    return deep_get(dictionary.get(keys[0]), keys[1:], default)


def _style_cells(cells: Iterable[Cell], style: Optional[Dict]) -> None:
    if style is None:
        # Do nothing
        return

    # Font: blindly assume JSON schema matches Font.__init__() parameters (reasonable enough)
    font_args = style.get("font")
    try:
        font = Font(**font_args) if font_args else None
    except (TypeError, ValueError) as err:
        raise ValueError("Invalid font specification", font_args) from err
    # Fill: the only relevant thing is the fill color
    fill_color = deep_get(style, ["fill", "color"])
    try:
        fill = PatternFill(start_color=fill_color, fill_type="solid") if fill_color else None
    except (TypeError, ValueError) as err:
        raise ValueError("Invalid fill color", fill_color) from err
    # Alignment: blindly assume JSON schema matches Alignment.__init__() params (reasonable enough)
    alignment_args = style.get("alignment")
    try:
        alignment = Alignment(**alignment_args,) if alignment_args else None
    except (TypeError, ValueError) as err:
        raise ValueError("Invalid alignment specification", alignment_args) from err

    for cell in cells:
        # Code inspection complains that Cell attributes are read-only, but mutating them is,
        # in fact, the correct, documented way of applying styles to cells.
        # Ref: https://openpyxl.readthedocs.io/en/stable/styles.html#cell-styles
        if font is not None:
            cell.font = font  # noqa
        if fill is not None:
            cell.fill = fill  # noqa
        if alignment is not None:
            cell.alignment = alignment  # noqa


def _style_tables_in_worksheet(
    ws: OpenpyxlWorksheet,
    table_dimensions: List[Tuple[int, int, bool]],
    styles: Dict,
    sep_lines: int,
) -> None:
    num_header_rows = 2
    num_name_unit_rows = 2

    rows = [row for row in ws.iter_rows()]
    i_start = 0

    # Loop through tables
    for i, (num_rows, num_cols, transposed) in enumerate(table_dimensions):
        # Figure out on what rows this table's various parts are located
        true_num_cols = num_cols
        true_num_rows = num_rows + num_name_unit_rows
        if transposed:
            # By definition, swap understanding of rows and columns
            true_num_cols, true_num_rows = true_num_rows, true_num_cols
        table_rows = [
            r[0:true_num_cols] for r in rows[i_start : i_start + true_num_rows + num_header_rows]
        ]

        table_name_cells = table_rows[0]
        destination_cells = table_rows[1]
        if transposed:
            column_name_cells = [t[0] for t in table_rows[2:]]
            column_unit_cells = [t[1] for t in table_rows[2:]]
            value_cells = [t[2:] for t in table_rows[2:]]
        else:
            column_name_cells = table_rows[2]
            column_unit_cells = table_rows[3]
            value_cells = table_rows[4:]

        # Apply all the styles
        style_spec_names = [
            (table_name_cells, "table_name"),
            (destination_cells, "destinations"),
            (column_name_cells, "column_names"),
            (column_unit_cells, "units"),
            (chain.from_iterable(value_cells), "values"),
        ]
        for cells, style_spec_name in style_spec_names:
            try:
                _style_cells(cells, styles.get(style_spec_name))
            except ValueError as err:
                raise ValueError(f"Invalid style specification for '{style_spec_name}'") from err

        # Special default case for transposed tables: center values and units
        if transposed:
            centered = {"alignment": {"horizontal": "center"}}
            if not deep_get(styles, ["units", "alignment", "horizontal"]):
                _style_cells(column_unit_cells, centered)
            if not deep_get(styles, ["values", "alignment", "horizontal"]):
                _style_cells(chain.from_iterable(value_cells), centered)

        i_start += true_num_rows + num_header_rows + sep_lines

    # Widen columns
    max_num_cols = 0
    for rows, cols, transposed in table_dimensions:
        true_num_cols = rows if transposed else cols
        max_num_cols = max(max_num_cols, true_num_cols)
    for i_column in [get_column_letter(i + 1) for i in range(max_num_cols + 1)]:
        ws.column_dimensions[i_column].width = 20
