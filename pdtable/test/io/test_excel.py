import io
import datetime
import tempfile
from pathlib import Path

import pytest
import pandas as pd
import openpyxl

try:
    from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet
except ImportError:
    # openpyxl < 2.6
    from openpyxl.worksheet import Worksheet as OpenpyxlWorksheet  # noqa: F401

from pdtable import Table, BlockType, write_excel, read_excel
from pdtable.io.excel import ExcelWriteBackend
from pdtable.io._excel_openpyxl import _append_table_to_openpyxl_worksheet, deep_get


def test_deep_get():
    d = {'meta': {'status': 'OK', 'status_code': 200}}
    assert deep_get(d, ['meta', 'status_code']) == 200
    assert deep_get(d, ['garbage', 'status_code']) is None
    assert deep_get(d, ['meta', 'garbage'], default='-') == '-'


def test__append_table_to_openpyxl_worksheet(places_table):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Act
    _append_table_to_openpyxl_worksheet(places_table, ws, sep_lines=1)

    # Assert worksheet looks as expected:
    # table header by row
    assert ws["A1"].value == "**foo"
    assert ws["A2"].value == "all"
    assert [ws.cell(3, c).value for c in range(1, 5)] == ["place", "distance", "ETA", "is_hot"]
    assert [ws.cell(4, c).value for c in range(1, 5)] == ["text", "km", "datetime", "onoff"]

    # table data by column
    assert [ws.cell(r, 1).value for r in range(5, 9)] == ["home", "work", "beach", "wonderland"]
    assert [ws.cell(r, 2).value for r in range(5, 9)] == [0, 1, 2, "-"]
    assert [ws.cell(r, 3).value for r in range(5, 9)] == list(
        pd.to_datetime(["2020-08-04 08:00", "2020-08-04 09:00", "2020-08-04 17:00"])
    ) + ["-"]
    assert [ws.cell(r, 4).value for r in range(5, 9)] == [1, 0, 1, 0]


@pytest.mark.parametrize("backend", list(ExcelWriteBackend))
def test_write_excel(tmp_path, places_table, backend):
    # This one is transposed
    t2 = Table(name="bar")
    t2.add_column("digit", [1, 6, 42], "-")
    t2.add_column("spelling", ["one", "six", "forty-two"], "text")
    t2.metadata.transposed = True

    # Write tables to workbook, save, and re-load
    out_path = tmp_path / "foo.xlsx"
    write_excel([places_table, t2], out_path, backend=backend)
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # First table is written as expected
    # - table header by row
    assert ws["A1"].value == "**foo"
    assert ws["A2"].value == "all"
    assert [ws.cell(3, c).value for c in range(1, 5)] == ["place", "distance", "ETA", "is_hot"]
    assert [ws.cell(4, c).value for c in range(1, 5)] == ["text", "km", "datetime", "onoff"]

    # - table data by column
    assert [ws.cell(r, 1).value for r in range(5, 9)] == ["home", "work", "beach", "wonderland"]
    assert [ws.cell(r, 2).value for r in range(5, 9)] == [0, 1, 2, "-"]
    for r, d in zip(
        range(5, 8), pd.to_datetime(["2020-08-04 08:00", "2020-08-04 09:00", "2020-08-04 17:00"])
    ):
        # workaround openpyxl bug: https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1493
        # openpyxl adds a spurious microsecond to some datetimes.
        assert abs(ws.cell(r, 3).value - d) <= datetime.timedelta(microseconds=1)
    assert ws.cell(8, 3).value == "-"
    assert [ws.cell(r, 4).value for r in range(5, 9)] == [1, 0, 1, 0]

    # Second table is there as well
    assert ws["A10"].value == "**bar*"
    assert ws["A11"].value == "all"
    # column headers (transposed)
    assert [ws.cell(r, 1).value for r in range(12, 14)] == ["digit", "spelling"]
    assert [ws.cell(r, 2).value for r in range(12, 14)] == ["-", "text"]
    # column values (transposed)
    assert [ws.cell(12, c).value for c in range(3, 6)] == [1, 6, 42]
    assert [ws.cell(13, c).value for c in range(3, 6)] == ["one", "six", "forty-two"]

    # Teardown
    out_path.unlink()


@pytest.mark.parametrize("backend", list(ExcelWriteBackend))
def test_write_excel__multiple_sheets(tmp_path, places_table, backend):
    """write_excel() can write tables to multiple sheets in a workbook"""

    # Make a couple of tables
    t2 = Table(name="bar")
    t2.add_column("digit", [1, 6, 42], "-")
    t2.add_column("spelling", ["one", "six", "forty-two"], "text")

    # Write tables to workbook, save, and re-load
    out_path = tmp_path / "foo.xlsx"
    write_excel({"sheet_one": [places_table, t2], "sheet_two": t2}, out_path, backend=backend)
    wb = openpyxl.load_workbook(out_path)

    # Workbook has the expected sheets
    assert len(wb.worksheets) == 2
    assert wb.sheetnames == ["sheet_one", "sheet_two"]
    # First sheet contains the expected tables
    assert wb.worksheets[0]["A1"].value == "**foo"
    assert wb.worksheets[0]["A10"].value == "**bar"
    # Second sheet contains the expected tables
    assert wb.worksheets[1]["A1"].value == "**bar"
    # Table details are tested elsewhere.

    # Teardown
    out_path.unlink()


@pytest.mark.parametrize("backend", list(ExcelWriteBackend))
def test_write_excel__style(tmp_path, places_table, backend):
    # Make a couple of tables
    # This one is transposed
    t2 = Table(name="bar")
    t2.add_column("digit", [1, 6, 42], "-")
    t2.add_column("spelling", ["one", "six", "forty-two"], "text")
    t2.metadata.transposed = True

    # This one is also transposed
    t3 = Table(name="bas")
    t3.add_column("digit", [1, 6, 42], "-")
    t3.add_column("spelling", ["one", "six", "forty-two"], "text")
    t3.metadata.transposed = True

    # Write tables to workbook, save, and re-load
    out_path = tmp_path / "foo.xlsx"
    write_excel([places_table, t2, t3], out_path, styles=True, backend=backend)
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # First table is written as expected
    # - table header by row
    assert ws["A1"].value == "**foo"
    assert ws["A2"].value == "all"
    assert [ws.cell(3, c).value for c in range(1, 5)] == ["place", "distance", "ETA", "is_hot"]
    assert [ws.cell(4, c).value for c in range(1, 5)] == ["text", "km", "datetime", "onoff"]

    # Check table formatting
    assert ws["A1"].fill.fill_type == "solid"
    assert ws["A1"].fill.start_color.value[2:] == "D9D9D9"  # xlsxwrite prepends hex code with FF
    assert ws["A1"].font.color.value[2:] == "1F4E78"
    assert ws["A1"].font.bold is True

    assert ws["A2"].fill.fill_type == "solid"
    assert ws["A2"].fill.start_color.value[2:] == "D9D9D9"
    assert ws["A2"].font.color.value[2:] == "808080"
    assert ws["A2"].font.bold is True

    assert [ws.cell(3, c).fill.fill_type for c in range(1, 5)] == ["solid"] * 4
    assert [ws.cell(3, c).fill.start_color.value[2:] for c in range(1, 5)] == ["F2F2F2"] * 4
    assert [ws.cell(3, c).font.bold for c in range(1, 5)] == [True] * 4

    assert [ws.cell(4, c).fill.fill_type for c in range(1, 5)] == ["solid"] * 4
    assert [ws.cell(4, c).fill.start_color.value[2:] for c in range(1, 5)] == ["F2F2F2"] * 4
    assert [ws.cell(4, c).font.bold for c in range(1, 5)] == [False] * 4

    # - table data by column
    assert [ws.cell(r, 1).value for r in range(5, 9)] == ["home", "work", "beach", "wonderland"]
    assert [ws.cell(r, 1).fill.fill_type for r in range(5, 9)] == [None] * 4
    assert [ws.cell(r, 1).font.bold for r in range(5, 9)] == [False] * 4

    assert [ws.cell(r, 2).value for r in range(5, 9)] == [0, 1, 2, "-"]
    assert [ws.cell(r, 2).fill.fill_type for r in range(5, 9)] == [None] * 4
    assert [ws.cell(r, 2).font.bold for r in range(5, 9)] == [False] * 4
    for r, d in zip(
        range(5, 8), pd.to_datetime(["2020-08-04 08:00", "2020-08-04 09:00", "2020-08-04 17:00"])
    ):
        # workaround openpyxl bug: https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1493
        # openpyxl adds a spurious microsecond to some datetimes.
        assert abs(ws.cell(r, 3).value - d) <= datetime.timedelta(microseconds=1)
        assert ws.cell(r, 3).fill.fill_type is None
        assert ws.cell(r, 3).font.bold is False
    assert ws.cell(8, 3).value == "-"
    assert [ws.cell(r, 4).value for r in range(5, 9)] == [1, 0, 1, 0]
    assert [ws.cell(r, 4).fill.fill_type for r in range(5, 9)] == [None] * 4
    assert [ws.cell(r, 4).font.bold for r in range(5, 9)] == [False] * 4

    # Second table is there as well
    assert ws["A10"].value == "**bar*"
    assert ws["A10"].fill.fill_type == "solid"
    assert ws["A10"].fill.start_color.value[2:] == "D9D9D9"
    assert ws["A10"].font.color.value[2:] == "1F4E78"
    assert ws["A10"].font.bold is True

    assert ws["A11"].value == "all"
    assert ws["A11"].fill.fill_type == "solid"
    assert ws["A11"].fill.start_color.value[2:] == "D9D9D9"
    assert ws["A11"].font.color.value[2:] == "808080"
    assert ws["A11"].font.bold is True

    # column headers (transposed)
    assert [ws.cell(r, 1).value for r in range(12, 14)] == ["digit", "spelling"]
    assert [ws.cell(r, 1).fill.fill_type for r in range(12, 14)] == ["solid"] * 2
    assert [ws.cell(r, 1).fill.start_color.value[2:] for r in range(12, 14)] == ["F2F2F2"] * 2
    assert [ws.cell(r, 1).font.bold for r in range(12, 14)] == [True] * 2

    assert [ws.cell(r, 2).value for r in range(12, 14)] == ["-", "text"]
    assert [ws.cell(r, 2).fill.fill_type for r in range(12, 14)] == ["solid"] * 2
    assert [ws.cell(r, 2).fill.start_color.value[2:] for r in range(12, 14)] == ["F2F2F2"] * 2
    assert [ws.cell(r, 2).font.bold for r in range(12, 14)] == [False] * 2

    # column values (transposed)
    assert [ws.cell(12, c).value for c in range(3, 6)] == [1, 6, 42]
    assert [ws.cell(12, c).fill.fill_type for c in range(3, 6)] == [None] * 3
    assert [ws.cell(12, c).font.bold for c in range(3, 6)] == [False] * 3

    assert [ws.cell(13, c).value for c in range(3, 6)] == ["one", "six", "forty-two"]
    assert [ws.cell(13, c).fill.fill_type for c in range(3, 6)] == [None] * 3
    assert [ws.cell(13, c).font.bold for c in range(3, 6)] == [False] * 3

    # Third table is there as well
    assert ws["A15"].value == "**bas*"
    assert ws["A15"].fill.fill_type == "solid"
    assert ws["A15"].fill.start_color.value[2:] == "D9D9D9"
    assert ws["A15"].font.color.value[2:] == "1F4E78"
    assert ws["A15"].font.bold is True

    assert ws["A16"].value == "all"
    assert ws["A16"].fill.fill_type == "solid"
    assert ws["A16"].fill.start_color.value[2:] == "D9D9D9"
    assert ws["A16"].font.color.value[2:] == "808080"
    assert ws["A16"].font.bold is True

    # Teardown
    out_path.unlink()


@pytest.mark.parametrize("backend", list(ExcelWriteBackend))
def test_write_excel__custom_style(tmp_path, backend):
    # Make a table
    t = Table(name="foo")
    t["place"] = ["home", "work", "beach", "wonderland"]
    t.add_column("distance", list(range(3)) + [float("nan")], "km")
    nc = len(t.column_names)
    nr = len(t.df)

    # Make a style specification as a JSON-like data structure
    style_spec = {
        "table_name": {
            "font": {
                "name": "Times New Roman",
                "size": 24,
                "color": "FF0000",   # RGB hex color code
                "bold": True,
            },
            "fill": {
                "color": "AAAAAA",
            },
        },
        "destinations": {
            "font": {
                "italic": True,
                "color": "0000FF",
            },
            "fill": {
                "color": "888888",
            },
        },
        "column_names": {
            "font": {
                "color": "444400",
                "bold": True,
            },
            "fill": {
                "color": "777777",
            },
        },
        "units": {
            "font": {
                "color": "440044",
            },  # --------------------- fill unspecified, leave untouched
        },
        "values": {
            "alignment": {
                "horizontal": "left",
            },
            "fill": {
                "color": "EEEEEE",
            },  # --------------------- font unspecified, leave untouched
        },
    }

    # Write tables to workbook, save, and re-load
    out_path = tmp_path / "foo_custom_style.xlsx"
    write_excel([t], out_path, styles=style_spec, backend=backend)
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # Check table formatting
    # table name
    assert ws["A1"].fill.fill_type == "solid"
    assert ws["A1"].fill.start_color.value[2:] == "AAAAAA"
    assert ws["A1"].font.size == 24
    assert ws["A1"].font.color.value[2:] == "FF0000"
    assert ws["A1"].font.name == "Times New Roman"
    assert ws["A1"].font.bold is True

    # destinations
    assert ws["A2"].fill.fill_type == "solid"
    assert ws["A2"].fill.start_color.value[2:] == "888888"
    assert ws["A2"].font.color.value[2:] == "0000FF"
    assert ws["A2"].font.bold is False
    assert ws["A2"].font.italic is True

    # column names
    assert [ws.cell(3, c).fill.fill_type for c in range(1, nc+1)] == ["solid"] * nc
    assert [ws.cell(3, c).fill.start_color.value[2:] for c in range(1, nc+1)] == ["777777"] * nc
    assert [ws.cell(3, c).font.color.value[2:] for c in range(1, nc+1)] == ["444400"] * nc
    assert [ws.cell(3, c).font.bold for c in range(1, nc+1)] == [True] * nc

    # column units
    assert [ws.cell(4, c).fill.fill_type for c in range(1, nc+1)] == [None] * nc  # left as default
    assert [ws.cell(4, c).font.color.value[2:] for c in range(1, nc+1)] == ["440044"] * nc
    assert [ws.cell(4, c).font.bold for c in range(1, nc+1)] == [False] * nc
    assert [ws.cell(4, c).alignment.horizontal for c in range(1, nc+1)] == [None] * nc

    # values
    assert [[ws.cell(4 + r, c).fill.fill_type for c in range(1, nc + 1)] for r in range(1, nr + 1)] == [["solid"] * nc] * nr
    assert [[ws.cell(4 + r, c).fill.start_color.value[2:] for c in range(1, nc + 1)] for r in range(1, nr + 1)] == [["EEEEEE"] * nc] * nr
    assert [[ws.cell(4 + r, c).alignment.horizontal for c in range(1, nc + 1)] for r in range(1, nr + 1)] == [["left"] * nc] * nr


@pytest.mark.parametrize(
    "err_msg_match,style_spec",
    [
        ("Invalid.*table_name", {"table_name": {"font": {"size": "NOT_A_NUMBER!!!"}}}),
        ("Invalid.*table_name", {"table_name": {"font": {"color": "NOT_A_COLOR!!!"}}}),
        ("Invalid.*destinations", {"destinations": {"fill": {"color": "NOT_A_COLOR!!!"}}}),
        ("Invalid.*values", {"values": {"alignment": {"horizontal": "FOOBAR!!!"}}}),
    ]
)
def test_write_excel_openpyxl_raises_error_on_invalid_style_spec(tmp_path, err_msg_match, style_spec):
    # Make a table
    t = Table(name="foo")
    t["place"] = ["home", "work", "beach", "wonderland"]
    t.add_column("distance", list(range(3)) + [float("nan")], "km")

    # Invalid font size
    with pytest.raises(ValueError, match=err_msg_match):
        write_excel([t], tmp_path / "foo_invalid_style.xlsx", styles=style_spec)


@pytest.mark.parametrize("backend", list(ExcelWriteBackend))
def test_write_excel__transposed_table_units_and_values_are_centered_by_default(tmp_path, backend):
    # Make a table
    t = Table(name="foo")
    t["place"] = ["home", "work", "beach", "wonderland"]
    t.add_column("distance", list(range(3)) + [float("nan")], "km")
    t.metadata.transposed = True
    nc = len(t.column_names)
    nr = len(t.df)

    # DEFAULT ALIGNMENT: CENTER
    # Write tables to workbook with default styles, save, and re-load
    out_path = tmp_path / "foo_custom_style.xlsx"
    write_excel([t], out_path, styles=True, backend=backend)  # <<< Default style
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # column units and values are centered
    assert [ws.cell(2 + c, 2).alignment.horizontal for c in range(1, nc+1)] == ["center"] * nc
    assert [[ws.cell(2 + c, 2 + r).alignment.horizontal for c in range(1, nc + 1)] for r in range(1, nr + 1)] == [["center"] * nc] * nr

    # DEFAULT ALIGNMENT NOT APPLIED when custom alignment is specified
    # Write tables to workbook with custom alignment styles, save, and re-load
    out_path = tmp_path / "foo_custom_style.xlsx"
    left = {"alignment": {"horizontal": "left"}}
    write_excel([t], out_path, styles={"units": left, "values": left}, backend=backend)  # << Custom alignment
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # column units and values are not centered
    assert [ws.cell(2 + c, 2).alignment.horizontal for c in range(1, nc+1)] == ["left"] * nc
    assert [[ws.cell(2 + c, 2 + r).alignment.horizontal for c in range(1, nc + 1)] for r in range(1, nr + 1)] == [["left"] * nc] * nr


@pytest.mark.parametrize("backend", list(ExcelWriteBackend))
def test_write_excel__sep_lines(tmp_path, backend):
    # Make a couple of tables
    t = Table(name="foo")
    t["place"] = ["home", "work", "beach", "wonderland"]
    t.add_column("distance", list(range(3)) + [float("nan")], "km")
    t.add_column(
        "ETA",
        pd.to_datetime(["2020-08-04 08:00", "2020-08-04 09:00", "2020-08-04 17:00", pd.NaT]),
        "datetime",
    )
    t.add_column("is_hot", [True, False, True, False], "onoff")

    # This one is transposed
    t2 = Table(name="bar")
    t2.add_column("digit", [1, 6, 42], "-")
    t2.add_column("spelling", ["one", "six", "forty-two"], "text")
    t2.metadata.transposed = True

    # This one is also transposed
    t3 = Table(name="bas")
    t3.add_column("digit", [1, 6, 42], "-")
    t3.add_column("spelling", ["one", "six", "forty-two"], "text")
    t3.metadata.transposed = True

    # Write tables to workbook, save, and re-load
    out_path = tmp_path / "foo.xlsx"
    write_excel([t, t2, t3], out_path, sep_lines=2, backend=backend)
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # Tables start on the expected rows
    assert ws["A1"].value == "**foo"
    assert ws["A11"].value == "**bar*"
    assert ws["A17"].value == "**bas*"


@pytest.mark.parametrize("backend", list(ExcelWriteBackend))
def test_write_excel__empty_table(tmp_path, backend):
    t = Table(name="foo")
    t2 = Table(name="bar")
    t2.metadata.transposed = True
    out_path = tmp_path / "foo.xlsx"
    write_excel([t, t2], out_path, sep_lines=2, backend=backend)
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws["A1"].value == "**foo"
    assert ws["A7"].value == "**bar*"


@pytest.mark.parametrize("backend", list(ExcelWriteBackend))
def test_read_write_excel__round_trip_with_styles(tmp_path, backend):
    """Round-trip reading and writing and re-reading preserves tables"""
    from pdtable import TableBundle, read_excel
    fn_in = Path(__file__).parent / "input" / "foo.xlsx"
    bundle = TableBundle(read_excel(fn_in))
    out_path = tmp_path / "foo_styled.xlsx"
    # Doesn't crash on write
    write_excel(bundle, out_path, styles=True, backend=backend)
    # Re-read bundle is same as first one
    bundle2 = TableBundle(read_excel(out_path))
    for t, t2 in zip(bundle, bundle2):
        assert t.equals(t2)


def test_read_excel():

    # Prepare the expected tables.
    # Note: deliberately not testing datetime columns due to upstream bug in openpyxl:
    # timestamps are off by one microsecond.
    # https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1493

    t0 = Table(name="places_to_go")
    t0["place"] = ["home", "work", "beach", "wonderland"]
    t0.add_column("distance", list(range(3)) + [float("nan")], "km")
    t0.add_column("is_hot", [True, False, True, False], "onoff")

    t1 = Table(name="spelling_numbers")
    t1.add_column("number", [1, 6, 42], "-")
    t1.add_column("spelling", ["one", "six, as formula", "forty-two"], "text")

    t2 = Table(name="this_one_is_transposed")
    t2.add_column("diameter", [1.23], "cm")
    t2.add_column("melting_point", [273], "K")

    expected_tables = [t0, t1, t2]
    expected_transposed_flag = [False, False, True]

    # Read tables from file
    blocks = read_excel(Path(__file__).parent / "input" / "foo.xlsx")
    tables_read = [block for (block_type, block) in blocks if block_type == BlockType.TABLE]
    assert len(expected_tables) == len(tables_read)

    # Tables read are equal to the expected ones
    for te, tr, flag in zip(expected_tables, tables_read, expected_transposed_flag):
        assert te.equals(tr)
        assert tr.metadata.transposed == flag

    # test_read_excel__from_stream
    with open(Path(__file__).parent / "input" / "foo.xlsx", "rb") as fh:
        blocks = read_excel(fh)
        tables_read_stream = [
            block for (block_type, block) in blocks if block_type == BlockType.TABLE
        ]
        assert len(expected_tables) == len(tables_read_stream)

    # read from tempfile (#77)
    with tempfile.TemporaryFile() as f:
        write_excel(t0, f)
        f.seek(0)
        assert t0.equals(list(read_excel(f))[0][1])


def test_read_excel__applies_filter():

    # Make a filter
    def is_table_about_spelling(block_type: BlockType, block_name: str) -> bool:
        return block_type == BlockType.TABLE and "spelling" in block_name

    # Read blocks from file
    blocks = list(
        read_excel(Path(__file__).parent / "input" / "foo.xlsx", filter=is_table_about_spelling)
    )

    # Assert only that one table block was parsed
    assert len(blocks) == 1
    table: Table = blocks[0][1]
    assert table.name == "spelling_numbers"


def test__table_is_preserved_when_written_to_and_read_from_excel():
    table_write = Table(
            pd.DataFrame({
                "a": [1, 2, 3],
                "b": ["a", "b", "c"]
            }),
            name="test",
            destinations="a b c")

    with io.BytesIO() as s:
        write_excel(tables=table_write, to=s)
        s.seek(0)
        block = next(read_excel(s))
    table_read = block[1]

    assert table_read.equals(table_write)
    assert table_read.name == table_write.name
    assert table_read.column_names == table_write.column_names
    assert table_read.units == table_write.units
    assert table_read.destinations == table_write.destinations
