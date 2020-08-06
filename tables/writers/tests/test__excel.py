import io
from textwrap import dedent

import pandas as pd

from tables import Table
from .._excel import _append_table_to_openpyxl_worksheet

import openpyxl

try:
    from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet
except ImportError:
    # openpyxl < 2.6
    from openpyxl.worksheet import Worksheet as OpenpyxlWorksheet


def test__append_table_to_openpyxl_worksheet():
    t = Table(name="foo")
    t["place"] = ["home", "work", "beach", "wonderland"]
    t.add_column("distance", list(range(3)) + [float("nan")], "km")
    t.add_column(
        "ETA",
        pd.to_datetime(["2020-08-04 08:00", "2020-08-04 09:00", "2020-08-04 17:00", pd.NaT]),
        "datetime",
    )
    t.add_column("is_hot", [True, False, True, False], "onoff")

    wb = openpyxl.Workbook()
    ws = wb.active

    # This is the "act" part:
    _append_table_to_openpyxl_worksheet(t, ws)
    wb.save("killme.xlsx")

    # table header by row
    assert ws["A1"].value == "**foo"
    assert ws["A2"].value == "all"
    assert [ws.cell(3, c).value for c in range(1, 5)] == ["place", "distance", "ETA", "is_hot"]
    assert [ws.cell(4, c).value for c in range(1, 5)] == ["text", "km", "datetime", "onoff"]

    # table data by column
    assert [ws.cell(r, 1).value for r in range(5, 9)] == ["home", "work", "beach", "wonderland"]
    assert [ws.cell(r, 2).value for r in range(5, 9)] == [0, 1, 2, "-"]
    assert [ws.cell(r, 3).value for r in range(5, 9)] == [
        x for x in pd.to_datetime(["2020-08-04 08:00", "2020-08-04 09:00", "2020-08-04 17:00"])
    ] + ["-"]
    assert [ws.cell(r, 4).value for r in range(5, 9)] == [1, 0, 1, 0]
